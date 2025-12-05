"""Batch processing endpoints for handwriting synthesis."""

import os
import sys
import shutil
import tempfile
import time
import zipfile
import json
import uuid
from typing import List, Tuple, Dict, Any
from flask import Blueprint, jsonify, request, send_file, Response, stream_with_context
from flask_login import login_required

# Ensure project root is in sys.path
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from handwriting_synthesis.hand.Hand import Hand
from webapp.utils.generation_utils import parse_generation_params, generate_handwriting_to_file


# Create blueprint
batch_bp = Blueprint('batch', __name__)

# Initialize Hand model (shared across requests)
hand = Hand()

# Jobs directory for streaming batch processing
JOBS_ROOT = os.path.join(tempfile.gettempdir(), "writebot_jobs")
os.makedirs(JOBS_ROOT, exist_ok=True)


def _get_row_value(row: Dict[str, Any], key: str, default=None):
    """
    Get value from row dictionary, handling NaN values.

    Args:
        row: Dictionary of row data.
        key: Key to retrieve.
        default: Default value if key is missing or NaN.

    Returns:
        Value from row or default.
    """
    v = row.get(key)
    return default if (v is None or (isinstance(v, float) and str(v) == "nan")) else v


@batch_bp.route("/api/batch", methods=["POST"])
@login_required
def batch_generate():
    """
    Process batch CSV/XLSX upload and generate multiple handwriting samples.

    Accepts a file upload ('file') containing batch generation parameters.
    Returns a ZIP file containing the generated SVG files.

    Returns:
        Response object with ZIP content or JSON error.
    """
    from webapp.utils.auth_utils import log_activity, track_generation
    if "file" not in request.files:
        return jsonify({"error": "CSV/XLSX file is required under 'file' field"}), 400

    uploaded_file = request.files["file"]
    try:
        import pandas as pd
        filename = uploaded_file.filename.lower()
        if filename.endswith('.xlsx'):
            try:
                df = pd.read_excel(uploaded_file, sheet_name='Data', engine='openpyxl')
            except:
                uploaded_file.seek(0)
                df = pd.read_excel(uploaded_file, sheet_name=0, engine='openpyxl')
        elif filename.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        else:
            return jsonify({"error": "File must be CSV or XLSX format"}), 400
    except Exception:
        return jsonify({"error": "Failed to read file"}), 400

    # Defaults from query/body
    defaults = request.form.to_dict(flat=True)
    # Non-form JSON defaults if provided
    try:
        json_defaults = request.get_json(silent=True) or {}
        defaults.update(json_defaults)
    except Exception:
        pass

    tmp_dir = tempfile.mkdtemp(prefix="writebot_batch_")
    out_dir = os.path.join(tmp_dir, "out")
    os.makedirs(out_dir, exist_ok=True)

    generated_files: List[str] = []
    errors: List[Tuple[int, str]] = []

    for idx, row in df.fillna("").iterrows():
        row_dict = row.to_dict()
        try:
            # Merge row data with defaults
            merged_params = {**defaults, **{k: v for k, v in row_dict.items() if v not in (None, "", "nan")}}

            # Ensure we have text
            if not merged_params.get("text"):
                raise ValueError("Empty text")

            # Handle line breaks: convert literal \n to actual newlines
            if "text" in merged_params:
                merged_params["text"] = merged_params["text"].replace("\\n", "\n")

            # Get filename and ensure .svg extension
            filename = _get_row_value(row_dict, "filename", f"sample_{idx}.svg")
            if not filename.lower().endswith('.svg'):
                filename = filename + '.svg'
            out_path = os.path.join(out_dir, os.path.basename(filename))

            # Parse all generation parameters using shared utility
            params = parse_generation_params(merged_params, defaults)

            # Generate using shared utility
            generate_handwriting_to_file(hand, out_path, params)

            generated_files.append(out_path)
        except Exception as e:
            errors.append((idx, str(e)))

    # Package ZIP
    zip_path = os.path.join(tmp_dir, f"writebot_batch_{int(time.time())}.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in generated_files:
            zf.write(path, arcname=os.path.basename(path))

    # If there were errors, add a log file
    if errors:
        error_log = os.path.join(tmp_dir, "errors.txt")
        with open(error_log, "w", encoding="utf-8") as f:
            for idx, msg in errors:
                f.write(f"row {idx}: {msg}\n")
        with zipfile.ZipFile(zip_path, "a", zipfile.ZIP_DEFLATED) as zf:
            zf.write(error_log, arcname="errors.txt")

    # Log the batch generation
    log_activity('batch', f'Generated batch with {len(generated_files)} files ({len(errors)} errors)')
    track_generation(lines_count=len(generated_files), chars_count=0,
                     processing_time=0, is_batch=True)

    return send_file(zip_path, mimetype="application/zip", as_attachment=True, download_name=os.path.basename(zip_path))


@batch_bp.route("/api/template-csv", methods=["GET"])
@login_required
def template_csv():
    """Download a blank template CSV file for batch processing."""
    # Comprehensive CSV template with all available parameters (headers only)
    header = (
        "filename,text,"
        "page_size,page_width,page_height,orientation,units,"
        "margin_top,margin_right,margin_bottom,margin_left,line_height,empty_line_spacing,"
        "align,background,global_scale,legibility,x_stretch,denoise,auto_size,manual_size_scale,"
        "biases,styles,stroke_colors,stroke_widths,"
        "wrap_char_px,wrap_ratio,wrap_utilization,"
        "use_chunked,adaptive_chunking,adaptive_strategy,words_per_chunk,chunk_spacing,max_line_width\n"
    )
    return Response(header, mimetype="text/csv", headers={
        'Content-Disposition': 'attachment; filename=writebot_template.csv'
    })


@batch_bp.route("/api/template-xlsx", methods=["GET"])
@login_required
def template_xlsx():
    """
    Download a formatted XLSX template.

    Includes data validation, instructions, and an about page.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
    from io import BytesIO
    from webapp.models import PageSizePreset

    # Fetch page size presets from database
    page_sizes = PageSizePreset.query.filter_by(is_active=True).order_by(PageSizePreset.name).all()
    page_size_names = [ps.name for ps in page_sizes] + ["Custom"]
    page_size_list = ",".join(page_size_names)

    wb = Workbook()

    # Set workbook metadata
    wb.properties.creator = "WriteBot"
    wb.properties.lastModifiedBy = "WriteBot"
    wb.properties.title = "WriteBot Batch Template"
    wb.properties.subject = "Handwriting Synthesis Batch Processing Template"
    wb.properties.description = "Template for batch processing handwriting generation with WriteBot"
    wb.properties.company = "WriteBot"

    # ===== DATA SHEET =====
    ws_data = wb.active
    ws_data.title = "Data"

    # Define headers with tooltips
    headers_with_notes = [
        ("filename", "Output filename WITHOUT extension (e.g., 'example1' or 'my-document'). The .svg extension will be added automatically."),
        ("text", "The text to convert to handwriting. Use \\n for line breaks (e.g., 'Line 1\\nLine 2')."),
        ("page_size", f"Page size preset. Available: {', '.join(page_size_names[:5])}... Select from dropdown. Use 'Custom' to specify custom dimensions."),
        ("page_width", "Custom page width (only used when page_size is 'Custom'). Specify in the units you selected (mm or px)."),
        ("page_height", "Custom page height (only used when page_size is 'Custom'). Specify in the units you selected (mm or px)."),
        ("orientation", "Page orientation: portrait (vertical) or landscape (horizontal)."),
        ("units", "Measurement units: mm (millimeters) or px (pixels). Used for margins and dimensions."),
        ("margin_top", "Top margin. Specify in the units you selected (mm or px). Default: 20."),
        ("margin_right", "Right margin. Specify in the units you selected (mm or px). Default: 20."),
        ("margin_bottom", "Bottom margin. Specify in the units you selected (mm or px). Default: 20."),
        ("margin_left", "Left margin. Specify in the units you selected (mm or px). Default: 20."),
        ("line_height", "Vertical spacing between lines of text. Leave empty for automatic spacing based on text size."),
        ("empty_line_spacing", "Spacing when you use \\\\ to create blank lines. Leave empty for automatic."),
        ("align", "Text alignment: left, center, or right. Aligns text horizontally on the page."),
        ("background", "Background color for the page. Use color names (white, blue) or hex codes (#FFFFFF). Use 'transparent' for no background."),
        ("global_scale", "Scales the entire output. 1.0 = normal size, >1.0 = larger, <1.0 = smaller."),
        ("legibility", "Controls how clear and readable the handwriting is. Higher = more legible but less natural. Options: natural, normal, high."),
        ("x_stretch", "Stretches handwriting horizontally. 1.0 = normal, >1.0 = wider, <1.0 = narrower."),
        ("denoise", "Smooths out handwriting strokes for cleaner lines. Recommended to keep enabled (true)."),
        ("auto_size", "Automatically adjust text size to fit the page. Set to true to enable."),
        ("manual_size_scale", "Manual size scale override when auto_size is false. 1.0 = normal."),
        ("biases", "Controls writing style variance. Lower (0.5-0.7) = more random/natural, higher (0.8-1.0) = more consistent. Separate with | for per-line (e.g., 0.75|0.8|0.6)."),
        ("styles", "Use different handwriting styles for each line. Separate style numbers with | (e.g., 9|9|12)."),
        ("stroke_colors", "Set ink color for each line. Use color names (black, blue) or hex codes (#333). Separate with | (e.g., black|blue|red)."),
        ("stroke_widths", "Controls pen thickness for each line. Default is 2. Separate with | for different widths per line (e.g., 2|3|2)."),
        ("wrap_char_px", "Estimated character width in pixels for text wrapping calculations."),
        ("wrap_ratio", "Text wrapping ratio for line breaking calculations."),
        ("wrap_utilization", "Wrap utilization factor for optimizing line width usage."),
        ("use_chunked", "Enable chunked generation for better quality and longer lines. Recommended: true."),
        ("adaptive_chunking", "Enable adaptive chunking based on text structure (word boundaries, punctuation)."),
        ("adaptive_strategy", "Strategy for adaptive chunking: balanced, word_length, sentence, punctuation, or fixed."),
        ("words_per_chunk", "Number of words per chunk when using chunked generation (e.g., 3)."),
        ("chunk_spacing", "Spacing between chunks in pixels (e.g., 8.0)."),
        ("max_line_width", "Maximum line width in coordinate units (default: 800)."),
    ]

    headers = [h[0] for h in headers_with_notes]

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write and style headers, add comments
    for col_num, (header, note) in enumerate(headers_with_notes, 1):
        cell = ws_data.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

        # Add comment/note to header
        comment = Comment(note, "WriteBot")
        comment.width = 300
        comment.height = 100
        cell.comment = comment

    # Set column widths
    # A=filename, B=text, C=page_size, D=page_width, E=page_height, F=orientation, G=units,
    # H=margin_top, I=margin_right, J=margin_bottom, K=margin_left, L=line_height, M=empty_line_spacing,
    # N=align, O=background, P=global_scale, Q=legibility, R=x_stretch, S=denoise, T=auto_size, U=manual_size_scale,
    # V=biases, W=styles, X=stroke_colors, Y=stroke_widths, Z=wrap_char_px, AA=wrap_ratio, AB=wrap_utilization,
    # AC=use_chunked, AD=adaptive_chunking, AE=adaptive_strategy, AF=words_per_chunk, AG=chunk_spacing, AH=max_line_width
    col_widths = {
        'A': 20, 'B': 50, 'C': 15, 'D': 12, 'E': 12, 'F': 12, 'G': 10,
        'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 18,
        'N': 10, 'O': 12, 'P': 12, 'Q': 12, 'R': 12, 'S': 10,
        'T': 12, 'U': 18, 'V': 15, 'W': 15, 'X': 15, 'Y': 15,
        'Z': 15, 'AA': 15, 'AB': 18, 'AC': 15, 'AD': 18, 'AE': 18,
        'AF': 18, 'AG': 15, 'AH': 15
    }
    for col, width in col_widths.items():
        ws_data.column_dimensions[col].width = width

    # Freeze header row
    ws_data.freeze_panes = "A2"

    # Set column data types and number formats
    # Column mapping: A=filename, B=text, C=page_size, D=page_width, E=page_height, F=orientation, G=units,
    # H=margin_top, I=margin_right, J=margin_bottom, K=margin_left, L=line_height, M=empty_line_spacing,
    # N=align, O=background, P=global_scale, Q=legibility, R=x_stretch, S=denoise, T=auto_size, U=manual_size_scale,
    # V=biases, W=styles, X=stroke_colors, Y=stroke_widths, Z=wrap_char_px, AA=wrap_ratio, AB=wrap_utilization,
    # AC=use_chunked, AD=adaptive_chunking, AE=adaptive_strategy, AF=words_per_chunk, AG=chunk_spacing, AH=max_line_width

    # Numeric columns with decimal format
    numeric_cols_decimal = ['D', 'E', 'H', 'I', 'J', 'K', 'L', 'M', 'P', 'R', 'U', 'Z', 'AA', 'AB', 'AG', 'AH']
    # page_width, page_height, margin_top, margin_right, margin_bottom, margin_left, line_height, empty_line_spacing,
    # global_scale, x_stretch, manual_size_scale, wrap_char_px, wrap_ratio, wrap_utilization, chunk_spacing, max_line_width
    for col in numeric_cols_decimal:
        for row in range(2, 1001):
            cell = ws_data[f'{col}{row}']
            cell.number_format = numbers.FORMAT_NUMBER_00  # Two decimal places

    # Integer columns
    integer_cols = ['AF']  # words_per_chunk
    for col in integer_cols:
        for row in range(2, 1001):
            cell = ws_data[f'{col}{row}']
            cell.number_format = numbers.FORMAT_NUMBER

    # Text columns (explicitly set for clarity)
    text_cols = ['A', 'B', 'C', 'F', 'G', 'N', 'O', 'Q', 'S', 'T', 'V', 'W', 'X', 'Y', 'AC', 'AD', 'AE']
    # filename, text, page_size, orientation, units, align, background, legibility, denoise, auto_size,
    # biases, styles, stroke_colors, stroke_widths, use_chunked, adaptive_chunking, adaptive_strategy
    for col in text_cols:
        for row in range(2, 1001):
            cell = ws_data[f'{col}{row}']
            cell.number_format = numbers.FORMAT_TEXT

    # Add data validation dropdowns
    # Page size dropdown (dynamic from database)
    page_size_dv = DataValidation(type="list", formula1=f'"{page_size_list}"', allow_blank=True)
    page_size_dv.error = 'Please select from the dropdown'
    page_size_dv.errorTitle = 'Invalid Page Size'
    ws_data.add_data_validation(page_size_dv)
    page_size_dv.add(f'C2:C1000')

    # Orientation dropdown
    orientation_dv = DataValidation(type="list", formula1='"portrait,landscape"', allow_blank=True)
    orientation_dv.error = 'Please select portrait or landscape'
    orientation_dv.errorTitle = 'Invalid Orientation'
    ws_data.add_data_validation(orientation_dv)
    orientation_dv.add(f'F2:F1000')

    # Units dropdown
    units_dv = DataValidation(type="list", formula1='"mm,px"', allow_blank=True)
    units_dv.error = 'Please select mm or px'
    units_dv.errorTitle = 'Invalid Units'
    ws_data.add_data_validation(units_dv)
    units_dv.add(f'G2:G1000')

    # Align dropdown
    align_dv = DataValidation(type="list", formula1='"left,center,right"', allow_blank=True)
    align_dv.error = 'Please select left, center, or right'
    align_dv.errorTitle = 'Invalid Alignment'
    ws_data.add_data_validation(align_dv)
    align_dv.add(f'N2:N1000')

    # Legibility dropdown
    legibility_dv = DataValidation(type="list", formula1='"natural,normal,high"', allow_blank=True)
    legibility_dv.error = 'Please select natural, normal, or high'
    legibility_dv.errorTitle = 'Invalid Legibility'
    ws_data.add_data_validation(legibility_dv)
    legibility_dv.add(f'Q2:Q1000')

    # Boolean dropdowns (true/false)
    bool_dv = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
    bool_dv.error = 'Please select true or false'
    bool_dv.errorTitle = 'Invalid Boolean'
    ws_data.add_data_validation(bool_dv)
    bool_dv.add(f'S2:S1000')  # denoise
    bool_dv.add(f'T2:T1000')  # auto_size
    bool_dv.add(f'AC2:AC1000')  # use_chunked
    bool_dv.add(f'AD2:AD1000')  # adaptive_chunking

    # Adaptive strategy dropdown
    strategy_dv = DataValidation(type="list", formula1='"balanced,word_length,sentence,punctuation,fixed"', allow_blank=True)
    strategy_dv.error = 'Please select from the dropdown'
    strategy_dv.errorTitle = 'Invalid Strategy'
    ws_data.add_data_validation(strategy_dv)
    strategy_dv.add(f'AE2:AE1000')

    # Template is blank - users fill in their own data
    # For examples, download the sample file instead

    # ===== INSTRUCTIONS SHEET =====
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.column_dimensions['A'].width = 100

    instructions_title_font = Font(bold=True, size=16, color="4472C4")
    instructions_heading_font = Font(bold=True, size=12, color="2E5090")
    instructions_text_font = Font(size=11)

    instructions_content = [
        ("WriteBot Batch Processing - Instructions", instructions_title_font),
        ("", None),
        ("How to Use This Template:", instructions_heading_font),
        ("", None),
        ("1. Fill in the 'Data' sheet with your batch processing requests", instructions_text_font),
        ("   - Each row represents one file to generate", instructions_text_font),
        ("   - Use the dropdowns for fields with predefined values", instructions_text_font),
        ("   - Leave optional fields blank to use defaults", instructions_text_font),
        ("", None),
        ("2. Required Fields:", instructions_heading_font),
        ("   - filename: Output name WITHOUT extension (e.g., 'example1')", instructions_text_font),
        ("     Note: .svg extension is added automatically", instructions_text_font),
        ("   - text: The text to convert to handwriting", instructions_text_font),
        ("", None),
        ("3. Line Breaks in Text:", instructions_heading_font),
        ("   - Use \\n to create line breaks (e.g., 'Line 1\\nLine 2')", instructions_text_font),
        ("   - Use \\\\ on its own line to create blank space", instructions_text_font),
        ("", None),
        ("4. Common Fields:", instructions_heading_font),
        ("   - page_size: A4, A5, Letter, Legal, or Custom (for custom dimensions)", instructions_text_font),
        ("   - page_width/page_height: Required when page_size is 'Custom'", instructions_text_font),
        ("   - orientation: portrait or landscape", instructions_text_font),
        ("   - units: mm (millimeters) or px (pixels)", instructions_text_font),
        ("   - margin_top/right/bottom/left: Individual margins for each side", instructions_text_font),
        ("   - align: left, center, or right", instructions_text_font),
        ("   - legibility: natural, normal, or high", instructions_text_font),
        ("", None),
        ("5. Advanced Fields:", instructions_heading_font),
        ("   - biases: Style variance control (e.g., '0.75|0.8' for per-line)", instructions_text_font),
        ("   - styles: Per-line style numbers (e.g., '9|9|12')", instructions_text_font),
        ("   - stroke_colors: Per-line colors (e.g., 'black|blue|red')", instructions_text_font),
        ("   - use_chunked: Enable chunked generation for better quality", instructions_text_font),
        ("   - adaptive_strategy: balanced, word_length, sentence, punctuation, or fixed", instructions_text_font),
        ("", None),
        ("6. Upload & Process:", instructions_heading_font),
        ("   - Save this file as .xlsx format", instructions_text_font),
        ("   - Upload to WriteBot batch processing", instructions_text_font),
        ("   - Click 'Start Batch Processing'", instructions_text_font),
        ("   - Download your generated files when complete", instructions_text_font),
        ("", None),
        ("Tips:", instructions_heading_font),
        ("   - Download the sample XLSX file to see example configurations", instructions_text_font),
        ("   - Copy rows to create similar configurations", instructions_text_font),
        ("   - The processing log will be included in your download", instructions_text_font),
    ]

    for row_num, (text, font) in enumerate(instructions_content, 1):
        cell = ws_instructions.cell(row=row_num, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ===== ABOUT SHEET =====
    ws_about = wb.create_sheet("About")
    ws_about.column_dimensions['A'].width = 100

    about_content = [
        ("About WriteBot", Font(bold=True, size=16, color="4472C4")),
        ("", None),
        ("WriteBot is a powerful handwriting synthesis tool that converts text into realistic handwritten documents.", Font(size=11)),
        ("", None),
        ("Features:", Font(bold=True, size=12, color="2E5090")),
        ("   • Multiple handwriting styles", Font(size=11)),
        ("   • Customizable page layouts and margins", Font(size=11)),
        ("   • Advanced text wrapping and alignment", Font(size=11)),
        ("   • Chunked generation for long text", Font(size=11)),
        ("   • Batch processing with this Excel template", Font(size=11)),
        ("   • Live preview and SVG output", Font(size=11)),
        ("", None),
        ("Batch Processing Benefits:", Font(bold=True, size=12, color="2E5090")),
        ("   • Process multiple files at once", Font(size=11)),
        ("   • Consistent styling across documents", Font(size=11)),
        ("   • Detailed processing logs", Font(size=11)),
        ("   • Easy data management with Excel", Font(size=11)),
        ("", None),
        ("Need Help?", Font(bold=True, size=12, color="2E5090")),
        ("   • Check the 'Instructions' sheet for detailed guidance", Font(size=11)),
        ("   • Download the sample XLSX file to see example configurations", Font(size=11)),
        ("   • Visit the main WriteBot interface for single file generation", Font(size=11)),
        ("", None),
        ("Version: 1.0", Font(size=10, italic=True)),
        ("Generated by WriteBot Batch Template System", Font(size=10, italic=True)),
    ]

    for row_num, (text, font) in enumerate(about_content, 1):
        cell = ws_about.cell(row=row_num, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='writebot_template.xlsx'
    )


@batch_bp.route("/api/sample-xlsx", methods=["GET"])
@login_required
def sample_xlsx():
    """Download a sample XLSX file with example data for testing/learning."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    wb = Workbook()

    # Set workbook metadata
    wb.properties.creator = "WriteBot"
    wb.properties.lastModifiedBy = "WriteBot"
    wb.properties.title = "WriteBot Sample File"
    wb.properties.subject = "Example configurations for WriteBot batch processing"
    wb.properties.description = "Sample file with example data to demonstrate WriteBot capabilities"
    wb.properties.company = "WriteBot"

    # ===== DATA SHEET =====
    ws_data = wb.active
    ws_data.title = "Sample Data"

    # Define styling
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Headers (same as template)
    headers = [
        "filename", "text", "page_size", "page_width", "page_height", "orientation", "units",
        "margin_top", "margin_right", "margin_bottom", "margin_left", "line_height", "empty_line_spacing",
        "align", "background", "global_scale", "legibility", "x_stretch", "denoise", "auto_size", "manual_size_scale",
        "biases", "styles", "stroke_colors", "stroke_widths", "wrap_char_px", "wrap_ratio", "wrap_utilization",
        "use_chunked", "adaptive_chunking", "adaptive_strategy", "words_per_chunk", "chunk_spacing", "max_line_width"
    ]

    # Set header row
    for col_num, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Example data rows
    example_data = [
        ["example1", "The quick brown fox jumps over the lazy dog.", "A4", "", "", "portrait", "mm",
         "20", "20", "20", "20", "", "", "left", "white", "1.0", "normal", "1.0", "true", "true", "",
         "", "", "", "", "", "", "", "true", "true", "balanced", "3", "8.0", "800"],
        ["example2", "Hello World!\\nThis is a new line.", "Letter", "", "", "portrait", "mm",
         "15", "15", "15", "15", "", "", "center", "white", "1.0", "high", "1.0", "true", "true", "",
         "", "", "", "", "", "", "", "true", "false", "", "", "", ""],
        ["custom_size", "Custom page dimensions example.", "Custom", "200", "300", "portrait", "mm",
         "10", "10", "10", "10", "", "", "left", "white", "1.0", "normal", "1.0", "true", "true", "",
         "", "", "", "", "", "", "", "true", "true", "balanced", "3", "8.0", "800"],
    ]

    # Add example rows
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws_data.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-size columns
    for col in ws_data.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 40)
        ws_data.column_dimensions[column].width = adjusted_width

    # ===== README SHEET =====
    ws_readme = wb.create_sheet("README")
    ws_readme.column_dimensions['A'].width = 100

    readme_content = [
        ("WriteBot Sample File", Font(bold=True, size=16, color="4472C4")),
        ("", None),
        ("This file contains example configurations to help you get started with WriteBot batch processing.", Font(size=11)),
        ("", None),
        ("What's in this file:", Font(bold=True, size=12, color="2E5090")),
        ("   • 3 example rows with different configurations", Font(size=11)),
        ("   • Basic example: Simple text with standard A4 page", Font(size=11)),
        ("   • Multi-line example: Text with line breaks and center alignment", Font(size=11)),
        ("   • Custom size example: Using custom page dimensions", Font(size=11)),
        ("", None),
        ("How to use this sample:", Font(bold=True, size=12, color="2E5090")),
        ("   1. Upload this file to WriteBot batch processing", Font(size=11)),
        ("   2. Click 'Start Batch Processing' to generate the examples", Font(size=11)),
        ("   3. Download and review the generated files", Font(size=11)),
        ("   4. Modify the examples or add your own rows", Font(size=11)),
        ("", None),
        ("Next steps:", Font(bold=True, size=12, color="2E5090")),
        ("   • Download the blank template to create your own batch", Font(size=11)),
        ("   • Check the 'Instructions' sheet in the template for full documentation", Font(size=11)),
        ("   • Experiment with different settings to learn what they do", Font(size=11)),
        ("", None),
        ("Happy handwriting generation!", Font(size=11, italic=True)),
    ]

    for row_num, (text, font) in enumerate(readme_content, 1):
        cell = ws_readme.cell(row=row_num, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='writebot_sample.xlsx'
    )


def _sse(obj: Dict[str, Any]) -> str:
    """
    Format a dictionary as a Server-Sent Event (SSE) message.

    Args:
        obj: Dictionary to serialize.

    Returns:
        Formatted SSE string.
    """
    return f"data: {json.dumps(obj)}\n\n"


@batch_bp.route("/api/batch/stream", methods=["POST"])
@login_required
def batch_stream():
    """
    Process batch CSV/XLSX with streaming progress updates (Server-Sent Events).

    Yields progress updates as JSON objects over an event stream.
    Final event contains the download URL for the results.

    Returns:
        Streamed response.
    """
    # Debug: Print what we received
    print(f"DEBUG: request.files keys: {list(request.files.keys())}")
    print(f"DEBUG: request.form keys: {list(request.form.keys())}")

    if "file" not in request.files:
        error_msg = f"CSV/XLSX file is required under 'file' field. Received files: {list(request.files.keys())}"
        print(f"ERROR: {error_msg}")
        return jsonify({"error": error_msg}), 400

    uploaded_file = request.files["file"]
    print(f"DEBUG: File received: {uploaded_file.filename}")

    try:
        import pandas as pd
        # Detect file type and read accordingly
        filename = uploaded_file.filename.lower()
        if filename.endswith('.xlsx'):
            # Read XLSX file - try 'Data' sheet first, otherwise first sheet
            try:
                df = pd.read_excel(uploaded_file, sheet_name='Data', engine='openpyxl')
            except:
                uploaded_file.seek(0)  # Reset file pointer
                df = pd.read_excel(uploaded_file, sheet_name=0, engine='openpyxl')
        elif filename.endswith('.csv'):
            # Important: Don't use first column as index
            df = pd.read_csv(uploaded_file, index_col=False)
        else:
            return jsonify({"error": "File must be CSV or XLSX format"}), 400

        print(f"DEBUG: File parsed successfully. Rows: {len(df)}, Columns: {list(df.columns)}")
    except Exception as e:
        error_msg = f"Failed to read file: {e}"
        print(f"ERROR: {error_msg}")
        return jsonify({"error": error_msg}), 400

    # Get defaults from form, but filter out None/empty values
    defaults = {k: v for k, v in request.form.to_dict(flat=True).items() if v}
    print(f"DEBUG: Form defaults: {defaults}")

    job_id = str(uuid.uuid4())
    job_dir = os.path.join(JOBS_ROOT, job_id)
    os.makedirs(job_dir, exist_ok=True)
    out_dir = os.path.join(job_dir, "out")
    os.makedirs(out_dir, exist_ok=True)
    zip_path = os.path.join(job_dir, "results.zip")

    def gen():
        # Initialize processing log
        log_lines = []
        log_lines.append('=' * 70)
        log_lines.append('WriteBot Batch Processing Log')
        log_lines.append('=' * 70)
        log_lines.append(f'Job ID: {job_id}')
        log_lines.append(f'Started at: {time.strftime("%Y-%m-%d %H:%M:%S")}')
        log_lines.append(f'Total rows to process: {len(df)}')
        log_lines.append('=' * 70)
        log_lines.append('')

        # Start event
        yield _sse({"type": "start", "job_id": job_id, "total": int(len(df))})

        generated_files: List[str] = []
        errors: List[Tuple[int, str]] = []
        start_time = time.time()

        for row_num, row in df.fillna("").iterrows():
            row_dict = row.to_dict()
            try:
                # CSV row values should override defaults, not the other way around
                # Filter out empty/nan values from CSV
                csv_values = {k: v for k, v in row_dict.items() if v not in (None, "", "nan", "NaN") and pd.notna(v)}

                # Merge: defaults first, then CSV overrides
                merged_params = {**defaults, **csv_values}

                # Ensure we have text
                if not merged_params.get("text"):
                    raise ValueError("Empty text")

                # Handle line breaks: convert literal \n to actual newlines
                if "text" in merged_params:
                    merged_params["text"] = merged_params["text"].replace("\\n", "\n")

                # Get filename and ensure .svg extension
                filename = _get_row_value(row_dict, "filename", f"sample_{row_num}.svg")
                if not filename.lower().endswith('.svg'):
                    filename = filename + '.svg'
                out_path = os.path.join(out_dir, os.path.basename(filename))

                print(f"DEBUG: Processing row {row_num}: filename={filename}")
                print(f"DEBUG: Merged params: {merged_params}")

                # Parse all generation parameters using shared utility
                params = parse_generation_params(merged_params, defaults)

                # Generate using shared utility
                row_start = time.time()
                generate_handwriting_to_file(hand, out_path, params)
                row_time = time.time() - row_start

                generated_files.append(out_path)
                log_lines.append(f'[✓] Row {row_num}: {filename} - SUCCESS (took {row_time:.2f}s)')

                yield _sse({
                    "type": "row",
                    "status": "ok",
                    "row": int(row_num),
                    "file": filename,
                    "job_id": job_id,
                })
            except Exception as e:
                print(f"ERROR: Row {row_num} failed: {e}")
                errors.append((int(row_num), str(e)))
                log_lines.append(f'[✗] Row {row_num}: ERROR - {str(e)}')

                yield _sse({
                    "type": "row",
                    "status": "error",
                    "row": int(row_num),
                    "error": str(e),
                    "job_id": job_id,
                })

            yield _sse({"type": "progress", "completed": int(row_num) + 1, "total": int(len(df))})

        # Add completion summary to log
        total_time = time.time() - start_time
        success_count = len(generated_files)
        error_count = len(errors)
        success_rate = (success_count / len(df) * 100) if len(df) > 0 else 0

        log_lines.append('')
        log_lines.append('=' * 70)
        log_lines.append('Processing Complete')
        log_lines.append('=' * 70)
        log_lines.append(f'Completed at: {time.strftime("%Y-%m-%d %H:%M:%S")}')
        log_lines.append(f'Total time: {total_time:.2f}s')
        log_lines.append(f'Average time per file: {(total_time/len(df)):.2f}s')
        log_lines.append(f'Total processed: {len(df)}')
        log_lines.append(f'Successful: {success_count} ({success_rate:.1f}%)')
        log_lines.append(f'Errors: {error_count}')
        log_lines.append('=' * 70)

        if errors:
            log_lines.append('')
            log_lines.append('Error Details:')
            log_lines.append('-' * 70)
            for idx, msg in errors:
                log_lines.append(f'  Row {idx}: {msg}')

        # Write processing log to file
        log_path = os.path.join(job_dir, "processing_log.txt")
        with open(log_path, "w", encoding="utf-8") as f:
            f.write('\n'.join(log_lines))

        # Package zip
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            # Add all generated files
            for path in generated_files:
                zf.write(path, arcname=os.path.basename(path))

            # Add processing log
            zf.write(log_path, arcname="processing_log.txt")

        download_url = f"/api/batch/result/{job_id}"
        yield _sse({
            "type": "done",
            "job_id": job_id,
            "download": download_url,
            "total": int(len(df)),
            "success": int(len(generated_files)),
            "errors": int(len(errors)),
        })

    headers = {
        "Content-Type": "text/event-stream",
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
    }
    return Response(stream_with_context(gen()), headers=headers)


@batch_bp.route("/api/batch/result/<job_id>", methods=["GET"])
@login_required
def batch_result(job_id: str):
    """
    Download batch processing results (ZIP file).

    Args:
        job_id: The ID of the batch job.

    Returns:
        File download response.
    """
    job_dir = os.path.normpath(os.path.join(JOBS_ROOT, job_id))
    # Make sure job_dir is still under JOBS_ROOT
    if not job_dir.startswith(os.path.abspath(JOBS_ROOT) + os.sep):
        return jsonify({"error": "Invalid job_id"}), 400
    zip_path = os.path.join(job_dir, "results.zip")
    if not os.path.isfile(zip_path):
        return jsonify({"error": "Result not found or expired"}), 404
    return send_file(zip_path, mimetype="application/zip", as_attachment=True, download_name=f"writebot_batch_{job_id}.zip")


@batch_bp.route("/api/batch/result/<job_id>/file/<path:filename>", methods=["GET"])
@login_required
def batch_result_file(job_id: str, filename: str):
    """
    Serve an individual generated file from a batch job for live preview.

    Args:
        job_id: The ID of the batch job.
        filename: The name of the file to retrieve.

    Returns:
        File content.
    """
    job_dir = os.path.normpath(os.path.join(JOBS_ROOT, job_id))
    # Make sure job_dir is still under JOBS_ROOT
    if not job_dir.startswith(os.path.abspath(JOBS_ROOT) + os.sep):
        return jsonify({"error": "Invalid job_id"}), 400
    out_dir = os.path.join(job_dir, "out")
    if not os.path.isdir(out_dir):
        return jsonify({"error": "Job not found or expired"}), 404
    # Prevent path traversal
    safe_name = os.path.basename(filename)
    file_path = os.path.join(out_dir, safe_name)
    if not os.path.isfile(file_path):
        return jsonify({"error": "File not found"}), 404
    # Guess mimetype by extension (default to SVG for .svg)
    mime = "image/svg+xml" if safe_name.lower().endswith(".svg") else None
    return send_file(file_path, mimetype=mime or "application/octet-stream", as_attachment=False, download_name=safe_name)
